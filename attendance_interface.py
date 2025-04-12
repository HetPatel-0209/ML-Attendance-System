import os
import tkinter as tk
from tkinter import ttk, messagebox
import datetime
import glob
import openpyxl
import pandas as pd

class AttendanceInterface:
    def __init__(self, root):
        self.root = root
        
        # Get current year
        self.current_year = datetime.datetime.now().year
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create form elements
        ttk.Label(main_frame, text="Attendance Data Selection", font=("Arial", 16)).grid(row=0, column=0, columnspan=2, pady=(0, 20), sticky="w")
        
        # Year (auto-filled with current year)
        ttk.Label(main_frame, text="Year:").grid(row=1, column=0, sticky="w", pady=5)
        self.year_var = tk.StringVar(value=str(self.current_year))
        self.year_entry = ttk.Entry(main_frame, textvariable=self.year_var, state="readonly")
        self.year_entry.grid(row=1, column=1, sticky="ew", pady=5)
        
        # Program selection
        ttk.Label(main_frame, text="Program:").grid(row=2, column=0, sticky="w", pady=5)
        self.program_var = tk.StringVar()
        self.program_combo = ttk.Combobox(main_frame, textvariable=self.program_var, state="readonly")
        self.program_combo.grid(row=2, column=1, sticky="ew", pady=5)
        
        # Semester selection
        ttk.Label(main_frame, text="Semester:").grid(row=3, column=0, sticky="w", pady=5)
        self.semester_var = tk.StringVar()
        self.semester_combo = ttk.Combobox(main_frame, textvariable=self.semester_var, state="readonly")
        self.semester_combo.grid(row=3, column=1, sticky="ew", pady=5)
        
        # Subject selection
        ttk.Label(main_frame, text="Subject:").grid(row=4, column=0, sticky="w", pady=5)
        self.subject_var = tk.StringVar()
        self.subject_combo = ttk.Combobox(main_frame, textvariable=self.subject_var, state="readonly")
        self.subject_combo.grid(row=4, column=1, sticky="ew", pady=5)
        
        # Apply button
        ttk.Button(main_frame, text="Apply Selection", command=self.apply_selection).grid(row=5, column=0, columnspan=2, pady=20)
        
        # Status label
        self.status_var = tk.StringVar()
        ttk.Label(main_frame, textvariable=self.status_var, foreground="blue").grid(row=6, column=0, columnspan=2, sticky="w")
        
        # Configure grid
        main_frame.columnconfigure(1, weight=1)
        
        # Bind events
        self.program_combo.bind("<<ComboboxSelected>>", self.on_program_selected)
        self.semester_combo.bind("<<ComboboxSelected>>", self.on_semester_selected)
        
        # Load initial data
        self.load_programs()
    
    def load_programs(self):
        """Load available programs from the directory structure"""
        base_path = f"data/attendanceData/{self.current_year}"
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        
        programs = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]
        self.program_combo['values'] = programs
        if programs:
            self.program_combo.current(0)
            self.on_program_selected(None)
    
    def on_program_selected(self, event):
        """Handle program selection"""
        selected_program = self.program_var.get()
        if not selected_program:
            return
        
        base_path = f"data/attendanceData/{self.current_year}/{selected_program}"
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        
        semesters = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]
        # Sort semesters numerically
        semesters.sort(key=lambda x: int(x.replace("sem", "")))
        self.semester_combo['values'] = semesters
        if semesters:
            self.semester_combo.current(0)
            self.on_semester_selected(None)
    
    def on_semester_selected(self, event):
        """Handle semester selection"""
        selected_program = self.program_var.get()
        selected_semester = self.semester_var.get()
        if not selected_program or not selected_semester:
            return
        
        base_path = f"data/attendanceData/{self.current_year}/{selected_program}/{selected_semester}"
        if not os.path.exists(base_path):
            os.makedirs(base_path, exist_ok=True)
        
        # Get all Excel files in the directory
        subject_files = glob.glob(os.path.join(base_path, "*.xlsx"))
        subjects = [os.path.splitext(os.path.basename(f))[0] for f in subject_files]
        
        if not subjects:
            # No subjects found, provide a text field to create a new one
            self.subject_combo['values'] = ["-- Create New Subject --"]
            self.subject_combo.current(0)
            self.subject_var.set("")  # Clear for user input
            self.subject_combo['state'] = 'normal'  # Allow typing
        else:
            self.subject_combo['values'] = subjects
            self.subject_combo.current(0)
            self.subject_combo['state'] = 'readonly'
    
    def apply_selection(self):
        """Apply the selected options and update app.py configuration"""
        year = self.year_var.get()
        program = self.program_var.get()
        semester = self.semester_var.get()
        subject = self.subject_var.get()
        
        if not program or not semester or not subject:
            messagebox.showerror("Error", "Please select all fields")
            return
        
        # Update app.py configuration
        self.update_app_configuration(year, program, semester, subject)
        
        self.status_var.set(f"Configuration applied for {subject} in {program} {semester}")
        
        # If this is a new subject, create the Excel file
        base_path = f"data/attendanceData/{year}/{program}/{semester}"
        subject_file = os.path.join(base_path, f"{subject}.xlsx")
        if not os.path.exists(subject_file):
            # Create new subject file
            self.create_new_subject_file(subject_file)
            messagebox.showinfo("Success", f"Created new subject file for {subject}")
    
    def update_app_configuration(self, year, program, semester, subject):
        """Update the app.py configuration with selected options"""
        config = {
            'year': year,
            'program': program,
            'semester': semester,
            'subject': subject,
            'filepath': f"data/attendanceData/{year}/{program}/{semester}/{subject}.xlsx"
        }
        
        # Save config to a file that app.py can read
        with open('attendance_config.txt', 'w') as f:
            for key, value in config.items():
                f.write(f"{key}={value}\n")
    
    def create_new_subject_file(self, filepath):
        """Create a new Excel file for a subject"""
        import openpyxl
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Create new workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = "Student Name"
        sheet.cell(row=1, column=2).value = "Student ID"
        
        # Add students from students.csv
        try:
            students_df = pd.read_csv('data/students.csv')
            for idx, row in students_df.iterrows():
                sheet.cell(row=idx+2, column=1).value = row['name']
                sheet.cell(row=idx+2, column=2).value = row['student_id']
        except (FileNotFoundError, pd.errors.EmptyDataError):
            pass  # No students to add
        
        # Save the workbook
        workbook.save(filepath)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Attendance Management Interface")
    root.geometry("600x400")
    app = AttendanceInterface(root)
    root.mainloop() 