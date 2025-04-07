from ctypes.wintypes import HFONT
import glob
import face_recognition
import numpy as np
from datetime import datetime
import cv2
import os
import openpyxl
import pandas as pd
import sys

# --- Function Definitions FIRST ---
def load_student_details():
    try:
        return pd.read_csv('data/students.csv')
    except (FileNotFoundError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=['student_id', 'name', 'course', 'email', 'phone'])

def get_student_details(name):
    df = load_student_details()
    student = df[df['name'] == name]
    return student.iloc[0].to_dict() if not student.empty else None

def load_attendance_config():
    config = {}
    try:
        with open('attendance_config.txt', 'r') as f:
            for line in f:
                key, value = line.strip().split('=', 1)
                config[key] = value
        return config
    except FileNotFoundError:
        current_year = datetime.now().year
        return {
            'year': str(current_year),
            'program': 'IT',
            'semester': 'sem6',
            'subject': 'ML',
            'filepath': f"data/attendanceData/{current_year}/IT/sem6/ML.xlsx"
        }

def create_or_open_attendance_workbook():
    config = load_attendance_config()
    year = config.get('year', str(datetime.now().year))
    program = config.get('program', 'IT')
    semester = config.get('semester', 'sem6')
    subject = config.get('subject', 'ML')
    filepath = config.get('filepath', f"data/attendanceData/{year}/{program}/{semester}/{subject}.xlsx")
    directory = os.path.dirname(filepath)
    os.makedirs(directory, exist_ok=True)
    try:
        workbook = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1).value = "Student Name"
        sheet.cell(row=1, column=2).value = "Student ID"
        workbook.save(filepath)
    sheet = workbook.active
    return workbook, sheet, filepath

def is_date_column_exists(sheet, date_str):
    for cell in sheet[1]:
        if cell.value == date_str:
            return cell.column
    return None

def mark_attendance(name):
    student_details = get_student_details(name)
    if not student_details:
        print(f"No details found for {name}")
        return
    workbook, sheet, filepath = create_or_open_attendance_workbook()
    current_date = datetime.now().strftime("%Y-%m-%d")
    student_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == name:
            student_row = row_idx
            break
    if student_row is None:
        student_row = sheet.max_row + 1
        sheet.cell(row=student_row, column=1).value = name
        sheet.cell(row=student_row, column=2).value = student_details.get('student_id', '')
    date_column = is_date_column_exists(sheet, current_date)
    if date_column is None:
        date_column = sheet.max_column + 1
        sheet.cell(row=1, column=date_column).value = current_date
    current_value = sheet.cell(row=student_row, column=date_column).value
    if current_value != 'P':
        sheet.cell(row=student_row, column=date_column).value = 'P'
        print(f"Attendance marked for {name} on {current_date}")
        workbook.save(filepath)
    else:
        print(f"Attendance already marked for {name} today")

def encoding1(images):
    encode = []
    for img in images:
        unk_encoding = face_recognition.face_encodings(img)[0]
        encode.append(unk_encoding)
    return encode

# --- Initialization and Main Logic AFTER functions ---
now = datetime.now()
dtString = now.strftime("%H:%M")
cap = cv2.VideoCapture(0)
running = True

def on_close():
    global running
    running = False
    if cap.isOpened():
        cap.release()
    cv2.destroyAllWindows()
    sys.exit()

cv2.namedWindow("Attendance System")
cv2.setWindowProperty("Attendance System", cv2.WND_PROP_TOPMOST, 1)

# Load student faces and encodings
images = []
names = []
path = os.path.join('faces', '*.*')
for file in glob.glob(path):
    image = cv2.imread(file)
    a = os.path.basename(file)
    b = os.path.splitext(a)[0]
    names.append(b)
    images.append(image)

encodelist = encoding1(images)
recently_recognized = {}

# Main loop
while running:
    ret, frame = cap.read()
    if not ret:
        print("Failed to grab frame")
        break

    frame1 = cv2.resize(frame, (0,0), None, 0.25, 0.25)
    face_locations = face_recognition.face_locations(frame1)
    curframe_encoding = face_recognition.face_encodings(frame1, face_locations)

    for encodeface, facelocation in zip(curframe_encoding, face_locations):
        distance = face_recognition.face_distance(encodelist, encodeface)
        match_index = np.argmin(distance)
        name = names[match_index]
        current_time = datetime.now().timestamp()
        if name not in recently_recognized or (current_time - recently_recognized[name]) > 60:
            mark_attendance(name)
            recently_recognized[name] = current_time
        student_details = get_student_details(name)
        x1, y1, x2, y2 = facelocation
        x1, y1, x2, y2 = x1*4, y1*4, x2*4, y2*4
        cv2.rectangle(frame, (y1, x1), (y2, x2), (0,0,255), 3)
        if student_details:
            y_offset = x2 + 30
            cv2.putText(frame, f"Name: {name}", (y1, y_offset), cv2.FONT_HERSHEY_COMPLEX, 0.6, (0,255,0), 1)
            if 'student_id' in student_details:
                y_offset += 25
                cv2.putText(frame, f"ID: {student_details['student_id']}", (y1, y_offset), cv2.FONT_HERSHEY_COMPLEX, 0.6, (0,255,0), 1)
            if 'course' in student_details:
                y_offset += 25
                cv2.putText(frame, f"Course: {student_details['course']}", (y1, y_offset), cv2.FONT_HERSHEY_COMPLEX, 0.6, (0,255,0), 1)
        else:
            cv2.putText(frame, name, (y2+6, x2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255,0,255), 2)

    config = load_attendance_config()
    config_text = f"{config.get('program', 'IT')} - {config.get('semester', 'sem6')} - {config.get('subject', 'ML')}"
    cv2.putText(frame, config_text, (10, 30), cv2.FONT_HERSHEY_COMPLEX, 0.6, (0, 255, 255), 1)
    cv2.imshow("Attendance System", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == 27 or key == ord('q'):
        on_close()
    if cv2.getWindowProperty("Attendance System", cv2.WND_PROP_VISIBLE) < 1:
        on_close()

on_close()