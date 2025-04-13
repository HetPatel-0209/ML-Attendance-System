import pandas as pd
import os
import pickle
import glob
import numpy as np
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import re
import openpyxl

def load_attendance_data():
    """Load all attendance files and format for processing."""
    attendance_data = []
    try:
        # Path to the attendance files
        base_path = "data/attendanceData"
        
        # Walk through all attendance files
        for year_dir in os.listdir(base_path):
            year_path = os.path.join(base_path, year_dir)
            if not os.path.isdir(year_path):
                continue
                
            for branch in os.listdir(year_path):
                branch_path = os.path.join(year_path, branch)
                if not os.path.isdir(branch_path):
                    continue
                    
                for sem in os.listdir(branch_path):
                    sem_path = os.path.join(branch_path, sem)
                    if not os.path.isdir(sem_path):
                        continue
                        
                    for subject_file in glob.glob(os.path.join(sem_path, "*.xlsx")):
                        try:
                            workbook = openpyxl.load_workbook(subject_file)
                            sheet = workbook.active
                            subject = os.path.splitext(os.path.basename(subject_file))[0]
                            
                            headers = [cell.value for cell in sheet[1]]
                            
                            for row_idx in range(2, sheet.max_row + 1):
                                student_name = sheet.cell(row=row_idx, column=1).value
                                student_id = sheet.cell(row=row_idx, column=2).value
                                
                                if not student_name:
                                    continue
                                    
                                # Process each date column
                                for col_idx in range(3, sheet.max_column + 1):
                                    date = headers[col_idx - 1]
                                    status = sheet.cell(row=row_idx, column=col_idx).value
                                    
                                    if status == 'P':
                                        attendance_data.append({
                                            "Name": student_name,
                                            "Student ID": student_id,
                                            "Date": date,
                                            "Status": "Present",
                                            "Year": year_dir,
                                            "Branch": branch,
                                            "Semester": sem,
                                            "Subject": subject
                                        })
                                    elif status == '' or status is None:
                                        attendance_data.append({
                                            "Name": student_name,
                                            "Student ID": student_id,
                                            "Date": date,
                                            "Status": "Absent",
                                            "Year": year_dir,
                                            "Branch": branch,
                                            "Semester": sem,
                                            "Subject": subject
                                        })
                        except Exception as e:
                            print(f"Error processing file {subject_file}: {e}")
                            continue
    except Exception as e:
        print(f"Error loading attendance data: {e}")
    
    if not attendance_data:
        return pd.DataFrame(columns=["Name", "Student ID", "Date", "Status", "Year", "Branch", "Semester", "Subject"])
    return pd.DataFrame(attendance_data)

def load_student_data():
    """Load student details from CSV."""
    try:
        return pd.read_csv('data/students.csv', dtype=str)
    except (FileNotFoundError, pd.errors.EmptyDataError):
        return pd.DataFrame(columns=['student_id', 'name', 'course', 'email', 'phone'])

def load_subject_data():
    """Load subject details from CSV."""
    try:
        return pd.read_csv('data/subjects.csv', dtype=str)
    except (FileNotFoundError, pd.errors.EmptyDataError):
        return pd.DataFrame()

def calculate_attendance_stats(attendance_df):
    """Calculate attendance statistics for each student."""
    if attendance_df.empty:
        return pd.DataFrame()
        
    stats = attendance_df.groupby(['Name', 'Subject']).agg({
        'Status': lambda x: (x == 'Present').mean()
    }).reset_index()
    
    stats = stats.rename(columns={'Status': 'Attendance_Percentage'})
    stats['Attendance_Percentage'] = stats['Attendance_Percentage'] * 100
    return stats

def generate_training_data(attendance_df, students_df, subjects_df):
    """Generate question-answer pairs for training."""
    qa_pairs = []
    
    # Basic validation
    if attendance_df.empty or students_df.empty:
        print("No data available to generate training examples")
        return qa_pairs
    
    # Calculate attendance statistics
    stats_df = calculate_attendance_stats(attendance_df)
    
    # 1. Student Information Queries
    for _, student in students_df.iterrows():
        # Student ID based queries
        student_id = student['student_id']
        name = student['name']
        
        # Different variations of enrollment/ID queries
        id_questions = [
            f"What is the name of student with enrollment {student_id}?",
            f"Who has enrollment number {student_id}?",
            f"Which student has ID {student_id}?",
            f"Find student with enrollment {student_id}",
            f"Get details of enrollment {student_id}",
            f"Student details for ID {student_id}"
        ]
        
        for question in id_questions:
            answer = f"The student with enrollment {student_id} is {name}"
            qa_pairs.append((question, answer))
        
        # Contact information queries
        contact_questions = [
            f"What is {name}'s phone number?",
            f"Give me contact details of {name}",
            f"What is the contact number of {name}?",
            f"How can I contact {name}?",
            f"What is {name}'s email?",
            f"What are {name}'s contact details?"
        ]
        
        contact_answer = f"Contact details for {name}: Email: {student['email']}, Phone: {student['phone']}"
        for question in contact_questions:
            qa_pairs.append((question, contact_answer))
            
        # Email specific queries
        email_questions = [
            f"What is {name}'s email address?",
            f"Tell me the email of {name}",
            f"What's the email for {name}?"
        ]
        
        email_answer = f"{name}'s email address is {student['email']}"
        for question in email_questions:
            qa_pairs.append((question, email_answer))
            
        # Course enrollment queries
        course_questions = [
            f"Which course is {name} enrolled in?",
            f"What is {name}'s course?",
            f"What branch is {name} in?"
        ]
        
        course_answer = f"{name} is enrolled in {student['course']}"
        for question in course_questions:
            qa_pairs.append((question, course_answer))
    
    # 2. Subject Code Queries
    for _, subject in subjects_df.iterrows():
        subject_name = subject['subject_name']
        subject_code = subject['subject_code']
        subject_abbrev = subject['subject_abbrevation']
        credits = subject['credits']
        branch = subject['branch']
        semester = subject['semester']
        subject_type = subject['subject_type']
        subject_type_abbreviation = subject['subject_type_abbreviation']
        total_classes = subject.get('total_classes', 'unknown')
        
        code_questions = [
            f"What is the subject code for {subject_name}?",
            f"What is the subject code of {subject_name}?",
            f"What is the code for {subject_name}?",
            f"Give me the code for {subject_name}",
            f"Subject code of {subject_name}?",
            f"What is the subject code for {subject_abbrev}?",
            f"What is the code for {subject_abbrev}?",
            f"How many {subject_type_abbreviation} subjects are there in {branch} branch?",
            f"How many {subject_type_abbreviation} subjects are there in {branch} branch in {semester} semester?"
            f"How many {subject_type} subjects are there in {branch} branch?",
            f"How many {subject_type} subjects are there in {branch} branch in {semester} semester?"
        ]
        
        code_answer = f"The subject code for {subject_name} ({subject_abbrev}) is {subject_code}"
        for question in code_questions:
            qa_pairs.append((question, code_answer))
            
        # Credit-related queries
        credit_questions = [
            f"How many credits does {subject_name} carry?",
            f"What is the credit value of {subject_name}?",
            f"How many credits is {subject_abbrev} worth?",
            f"Tell me the credits for {subject_name}"
        ]
        
        credit_answer = f"{subject_name} ({subject_abbrev}) carries {credits} credits"
        for question in credit_questions:
            qa_pairs.append((question, credit_answer))
            
        # Class count queries
        class_questions = [
            f"How many total classes are there for {subject_name}?",
            f"What is the total number of classes for {subject_abbrev}?",
            f"How many classes does {subject_name} have?"
        ]
        
        class_answer = f"{subject_name} ({subject_abbrev}) has {total_classes} total classes"
        for question in class_questions:
            qa_pairs.append((question, class_answer))
            
        # Subject type queries
        type_questions = [
            f"Is {subject_name} an elective?",
            f"What type of subject is {subject_name}?",
            f"Is {subject_abbrev} a regular subject or an elective?",
            f"Is {subject_name} a Professional Elective or Open Elective?"
        ]
        
        type_answer = f"{subject_name} ({subject_abbrev}) is a {subject_type} subject"
        for question in type_questions:
            qa_pairs.append((question, type_answer))
    
    # 3. Attendance Queries
    for name in students_df['name'].unique():
        student_attendance = attendance_df[attendance_df['Name'] == name]
        
        for subject in subjects_df['subject_name'].unique():
            subject_attendance = student_attendance[student_attendance['Subject'] == subject]
            
            if not subject_attendance.empty:
                # Present dates in subject
                present_dates = subject_attendance[subject_attendance['Status'] == 'Present']['Date'].tolist()
                present_dates_str = ', '.join(sorted(present_dates)) if present_dates else "no dates"
                
                question = f"On which dates was {name} present in {subject}?"
                answer = f"{name} was present in {subject} on: {present_dates_str}"
                qa_pairs.append((question, answer))
                
                # Absent dates in subject
                absent_dates = subject_attendance[subject_attendance['Status'] == 'Absent']['Date'].tolist()
                absent_dates_str = ', '.join(sorted(absent_dates)) if absent_dates else "no dates"
                
                question = f"On which dates was {name} absent in {subject}?"
                answer = f"{name} was absent in {subject} on: {absent_dates_str}"
                qa_pairs.append((question, answer))
                
                # Attendance on specific date
                for date in set(present_dates + absent_dates):
                    status = "present" if date in present_dates else "absent"
                    questions = [
                        f"Was {name} present in the {subject} class on {date}?",
                        f"Did {name} attend {subject} class on {date}?",
                        f"Show {name}'s attendance record for {subject} on {date}",
                        f"What was {name}'s attendance status for {subject} on {date}?"
                    ]
                    
                    answer = f"{name} was {status} in the {subject} class on {date}"
                    for q in questions:
                        qa_pairs.append((q, answer))
                
                # Attendance percentage in subject
                if not stats_df.empty:
                    subject_stats = stats_df[(stats_df['Name'] == name) & (stats_df['Subject'] == subject)]
                    if not subject_stats.empty:
                        percentage = subject_stats.iloc[0]['Attendance_Percentage']
                        question = f"What is {name}'s attendance in {subject}?"
                        answer = f"{name}'s attendance in {subject} is {percentage:.1f}%"
                        qa_pairs.append((question, answer))
                        
                        # Add more variations of the same question
                        question = f"Show me {name}'s attendance record in {subject}"
                        qa_pairs.append((question, answer))
                        
                        question = f"Calculate {name}'s total attendance percentage in {subject}"
                        qa_pairs.append((question, answer))
    
    # 4. Semester Subject Queries
    for branch in subjects_df['branch'].unique():
        for semester in subjects_df['semester'].unique():
            # Filter subjects for this branch and semester
            sem_subjects = subjects_df[(subjects_df['branch'] == branch) & 
                                      (subjects_df['semester'] == semester)]
            
            if not sem_subjects.empty:
                # All subjects in semester
                subject_list = ", ".join(sem_subjects['subject_name'].tolist())
                questions = [
                    f"Which subjects are offered in Semester {semester} for {branch.upper()}?",
                    f"List all subjects in Semester {semester} for {branch.upper()}",
                    f"What subjects are taught in Sem {semester} {branch.upper()}?"
                ]
                
                answer = f"Subjects offered in Semester {semester} for {branch.upper()}: {subject_list}"
                for q in questions:
                    qa_pairs.append((q, answer))
                
                # Elective subjects in semester
                pe_subjects = sem_subjects[sem_subjects['subject_type'] == 'Professional Elective']
                oe_subjects = sem_subjects[sem_subjects['subject_type'] == 'Open Elective']
                
                if not pe_subjects.empty:
                    pe_list = ", ".join(pe_subjects['subject_name'].tolist())
                    pe_questions = [
                        f"List all Professional Elective (PE) subjects in Semester {semester}",
                        f"What are the Professional Electives for Semester {semester}?",
                        f"Which PE subjects are available in Sem {semester}?"
                    ]
                    
                    pe_answer = f"Professional Elective subjects in Semester {semester}: {pe_list}"
                    for q in pe_questions:
                        qa_pairs.append((q, pe_answer))
                
                if not oe_subjects.empty:
                    oe_list = ", ".join(oe_subjects['subject_name'].tolist())
                    oe_count = len(oe_subjects)
                    
                    oe_questions = [
                        f"List all Open Elective (OE) subjects in Semester {semester}",
                        f"What are the Open Electives for Semester {semester}?",
                        f"How many Open Elective (OE) subjects are available in Semester {semester}?"
                    ]
                    
                    oe_answer = f"There are {oe_count} Open Elective subjects in Semester {semester}: {oe_list}"
                    for q in oe_questions:
                        qa_pairs.append((q, oe_answer))
                
                # Regular subjects
                reg_subjects = sem_subjects[sem_subjects['subject_type'] == 'Regular']
                if not reg_subjects.empty:
                    reg_list = ", ".join(reg_subjects['subject_name'].tolist())
                    reg_questions = [
                        f"Which subjects in Semester {semester} are labeled 'Regular'?",
                        f"List all Regular subjects in Sem {semester}"
                    ]
                    
                    reg_answer = f"Regular subjects in Semester {semester}: {reg_list}"
                    for q in reg_questions:
                        qa_pairs.append((q, reg_answer))
                
                # Credit distribution
                total_credits = sem_subjects['credits'].astype(int).sum()
                credit_questions = [
                    f"What is the credit distribution for Semester {semester} subjects?",
                    f"What is the total number of credits required to complete Semester {semester}?",
                    f"How many credits in total for Sem {semester}?"
                ]
                
                credits_by_subject = [f"{row['subject_name']} ({row['credits']} credits)" 
                                     for _, row in sem_subjects.iterrows()]
                credit_distrib = ", ".join(credits_by_subject)
                
                credit_answer = f"Total credits for Semester {semester}: {total_credits}. Distribution: {credit_distrib}"
                for q in credit_questions:
                    qa_pairs.append((q, credit_answer))
                
                # Zero-credit subjects
                zero_cred = sem_subjects[sem_subjects['credits'] == '0']
                if not zero_cred.empty:
                    zero_list = ", ".join(zero_cred['subject_name'].tolist())
                    zero_questions = [
                        f"Are there any zero-credit subjects in Semester {semester}?",
                        f"Which subjects in Sem {semester} have zero credits?",
                        f"List all zero-credit courses in Semester {semester}"
                    ]
                    
                    zero_answer = f"Zero-credit subjects in Semester {semester}: {zero_list}"
                    for q in zero_questions:
                        qa_pairs.append((q, zero_answer))
                else:
                    zero_questions = [
                        f"Are there any zero-credit subjects in Semester {semester}?",
                        f"Does Sem {semester} have any zero-credit courses?"
                    ]
                    
                    zero_answer = f"There are no zero-credit subjects in Semester {semester}"
                    for q in zero_questions:
                        qa_pairs.append((q, zero_answer))
    
    # 5. Adding more advanced queries - cross-referencing and analysis
    # Subject with highest number of classes
    if 'total_classes' in subjects_df.columns:
        for semester in subjects_df['semester'].unique():
            sem_subjects = subjects_df[subjects_df['semester'] == semester]
            if not sem_subjects.empty:
                sem_subjects.loc[:, 'total_classes'] = sem_subjects['total_classes'].astype(int)
                max_classes_subject = sem_subjects.loc[sem_subjects['total_classes'].idxmax()]
                
                questions = [
                    f"Which subject has the highest number of classes in Semester {semester}?",
                    f"What is the subject with most classes in Sem {semester}?"
                ]
                
                answer = f"The subject with the highest number of classes in Semester {semester} is {max_classes_subject['subject_name']} with {max_classes_subject['total_classes']} classes"
                for q in questions:
                    qa_pairs.append((q, answer))
    
    # Total number of electives across semesters
    pe_count = len(subjects_df[subjects_df['subject_type'] == 'Professional Elective'])
    oe_count = len(subjects_df[subjects_df['subject_type'] == 'Open Elective'])
    
    elective_questions = [
        "What is the total number of Professional Electives across all semesters?",
        "How many PE subjects are offered in total?",
        "How many Professional Elective courses are available?"
    ]
    
    elective_answer = f"There are {pe_count} Professional Elective subjects available across all semesters"
    for q in elective_questions:
        qa_pairs.append((q, elective_answer))
    
    oe_questions = [
        "What is the total number of Open Electives across all semesters?",
        "How many OE subjects are offered in total?",
        "How many Open Elective courses are available?"
    ]
    
    oe_answer = f"There are {oe_count} Open Elective subjects available across all semesters"
    for q in oe_questions:
        qa_pairs.append((q, oe_answer))
    
    return qa_pairs

def train_retrieval_model(qa_pairs):
    """Train a simple retrieval-based chatbot model."""
    if not qa_pairs:
        print("No training data available")
        return None
    
    questions = [pair[0] for pair in qa_pairs]
    answers = [pair[1] for pair in qa_pairs]
    
    # Create a TF-IDF vectorizer
    vectorizer = TfidfVectorizer()
    question_vectors = vectorizer.fit_transform(questions)
    
    # Save the model components
    os.makedirs("models", exist_ok=True)
    
    model = {
        'vectorizer': vectorizer,
        'question_vectors': question_vectors,
        'questions': questions,
        'answers': answers
    }
    
    with open('models/chatbot_model.pkl', 'wb') as f:
        pickle.dump(model, f)
    
    print("Chatbot model trained and saved successfully!")
    return model

def main():
    print("Loading attendance data...")
    attendance_df = load_attendance_data()
    
    print("Loading student data...")
    students_df = load_student_data()
    
    print("Loading subject data...")
    subjects_df = load_subject_data()
    
    print("Generating question-answer pairs...")
    qa_pairs = generate_training_data(attendance_df, students_df, subjects_df)
    
    print(f"Generated {len(qa_pairs)} question-answer pairs")
    
    print("Training the chatbot model...")
    train_retrieval_model(qa_pairs)

if __name__ == "__main__":
    main()
